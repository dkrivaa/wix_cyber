import os
import subprocess
import pandas as pd
import time

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from scenarioData import get_data
from simulation import run_simulation
from slides import make_slides


def format_excel_file(file_path):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Apply styles to the header (first row)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="right", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Set the width of the first column to 45
    ws.column_dimensions['A'].width = 45

    ws['A1'].alignment = Alignment(horizontal="left", vertical="center")

    # Set the width of all other columns to 11 and apply number format #,###
    for col in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col)  # Convert column number to column letter

        # Set the width of the column
        ws.column_dimensions[col_letter].width = 11

        # Apply the number format #,### to all cells in this column (starting from row 2)
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]  # Access the specific cell
            cell.number_format = '#,###'


    # Make specific rows bold
    bold_rows = [2, 7, 12, 17, 22, 32, ]
    bold_font = Font(bold=True)
    for row in bold_rows:
        for cell in ws[row]:
            cell.font = bold_font

    total_rows = [26, 39]
    for row in total_rows:
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill

    last_row = ws.max_row + 1
    ws[f'A{last_row}'] = 'Accumulated Gross Profit'
    ws[f'B{last_row}'] = ws[f'B{last_row - 1}'].value
    ws[f'B{last_row}'].number_format = '#,###'
    ws[f'A{last_row}'].font = header_font
    ws[f'B{last_row}'].font = header_font
    ws[f'A{last_row}'].fill = PatternFill(start_color="273f5c", end_color="273f5c", fill_type="solid")
    ws[f'B{last_row}'].fill = PatternFill(start_color="273f5c", end_color="273f5c", fill_type="solid")

    for col in range(3, ws.max_column + 1):
        col_letter = get_column_letter(col)
        prev_letter = get_column_letter(col - 1)
        ws[f'{col_letter}{last_row}'] = ws[f'{prev_letter}{last_row}'].value + ws[f'{col_letter}{last_row-1}'].value

        ws[f'{col_letter}{last_row}'].number_format = '#,###'
        ws[f'{col_letter}{last_row}'].font = header_font
        ws[f'{col_letter}{last_row}'].fill = PatternFill(start_color="273f5c", end_color="273f5c", fill_type="solid")


    # Save the formatted Excel file
    wb.save(file_path)


def remove_files():
    # Check if 'simulation.xlsx' exists in the repository and delete
    if os.path.exists('simulation.xlsx'):
        # Remove the file from the Git repository and stage the change
        subprocess.run(['git', 'rm', 'simulation.xlsx'], check=True)

        # Commit the deletion
        commit_message = 'Remove existing simulation.xlsx before generating a new one'
        subprocess.run(['git', 'commit', '-m', commit_message], check=True)

        # Push the deletion to the remote repository
        subprocess.run(['git', 'push'], check=True)

    # Check if 'simulation.pptx' exists in the repository and delete
    if os.path.exists('simulation.pptx'):
        # Remove the file from the Git repository and stage the change
        subprocess.run(['git', 'rm', 'simulation.pptx'], check=True)

        # Commit the deletion
        commit_message = 'Remove existing simulation.pptx before generating a new one'
        subprocess.run(['git', 'commit', '-m', commit_message], check=True)

        # Push the deletion to the remote repository
        subprocess.run(['git', 'push'], check=True)


def test():

    remove_files()
    # Wait for a few seconds to ensure the deletions are processed
    # time.sleep(10)

    # Get simulation data and run simulation
    data_dict = get_data()
    print(data_dict['scenarioName'])

    results = run_simulation(data_dict)

    df = pd.DataFrame(results, columns=[
        'month', 'customers', 'new customers', 'referred customers', 'lead customers',
        'existing customers', 'Risk assessment pakages sold', 'Risk assessment pakages sold to new customers',
        'Risk assessment pakages sold to referred customers', 'Risk assessment pakages sold to lead customers',
        'Risk assessment pakages sold to existing customers', 'SOC pakages sold',
        'SOC pakages sold to new customers', 'SOC pakages sold to referred customers',
        'SOC pakages sold to lead customers', 'SOC pakages sold to existing customers', 'Insurance packages sold',
        'Insurance packages sold to new customers', 'Insurance packages sold to referred customers',
        'Insurance packages sold to lead customers', 'Insurance packages sold to existing customers',
        'All packages sold', 'Income from risk assessment packages', 'Income from SOC packages', 'Income from insurance packages',
        'Total income', 'Admin staff', 'Tele staff', 'Sales staff', 'Cyber staff', 'Logistics staff',
        'Total staff', 'Labor cost', 'Risk assessment packages cost', 'SOC packages cost', 'Marketing cost',
        'General overhead', 'Legal & Accounting cost', 'Total cost', 'Gross profit', 'Acc. Gross Profit',
    ])

    # Make slides
    make_slides(df, data_dict, 'simulation.pptx')

    df = df.T
    df = df.reset_index()
    df.to_excel('simulation.xlsx', index=False, header=False)

    # Format Excel file
    format_excel_file('simulation.xlsx')

    # Save new EXCEL file with simulation results
    if os.path.exists('simulation.xlsx') or os.path.exists('simulation.pptx'):
        remove_files()
    else:
        # Add the Excel file to the Git staging area
        subprocess.run(['git', 'add', 'simulation.xlsx'], check=True)

        # Commit the Excel file with a message
        commit_message = 'Add Excel file generated by GitHub Actions workflow'
        subprocess.run(['git', 'commit', '-m', commit_message], check=True)

        # Push the changes back to the repository
        subprocess.run(['git', 'push'], check=True)

        # Save new POWERPOINT file with simulation results
        # Add the Powerpoint file to the Git staging area
        subprocess.run(['git', 'add', 'simulation.pptx'], check=True)

        # Commit the Powerpoint file with a message
        commit_message = 'Add Powerpoint file generated by GitHub Actions workflow'
        subprocess.run(['git', 'commit', '-m', commit_message], check=True)

        # Push the changes back to the repository
        subprocess.run(['git', 'push'], check=True)



    # time.sleep(30)


if __name__ == '__main__':
    test()



